# -*- coding: utf-8 -*-
"""
Conversor de extratos bancarios -> Excel (padrao Ello Solar)

Entrada : extrato bancario em PDF  OU  arquivo OFX
Saida   : .xlsx com duas abas -> "Debito" e "Credito"
          colunas: Data | Descricao | Valor | Obs

- OFX  : lido de forma 100% deterministica (sem IA, sem custo).
- PDF  : lido pela API da Anthropic (Claude), que funciona para qualquer
         layout de banco, inclusive extratos escaneados.
"""

import io
import os
import re
import json
import time
import base64
import hashlib
from datetime import datetime

import yaml
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ----------------------------------------------------------------------------
# Config geral
# ----------------------------------------------------------------------------
st.set_page_config(page_title="Conversor de extratos", page_icon="📄", layout="wide")

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
USERS_FILE = os.path.join(BASE_DIR, "users.yaml")

MODELO_CLAUDE       = "claude-sonnet-5"   # modelo usado para ler o PDF
CHUNK_PAGINAS_TEXTO = 6                   # PDF com texto: paginas por chamada a IA
CHUNK_PAGINAS_PDF   = 3                   # PDF escaneado (imagem): paginas por chamada
MAX_TOKENS_SAIDA    = 32000
TENTATIVAS_LOTE     = 3                   # re-tentativas por lote antes de desistir

CABECALHO = ["Data", "Descrição", "Valor", "Obs"]

# ============================================================================
# 1. AUTENTICACAO  (mesmo esquema simples usado nas outras ferramentas)
# ============================================================================
def carregar_usuarios():
    if not os.path.exists(USERS_FILE):
        return {}
    with open(USERS_FILE, "r", encoding="utf-8") as f:
        return (yaml.safe_load(f) or {}).get("usuarios", {})


def salvar_usuarios(u):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        yaml.safe_dump({"usuarios": u}, f, allow_unicode=True)


def criar_hash(s):
    return "sha256:" + hashlib.sha256(s.encode()).hexdigest()


def verificar_login(usuario, senha):
    usuario = (usuario or "").lower().strip()
    u = carregar_usuarios().get(usuario)
    if not u or not u.get("ativo", True):
        return False, None
    h = u.get("senha_hash", "")
    if h.startswith("sha256:") and h == criar_hash(senha):
        return True, u
    return False, None


def precisa_configurar_senha():
    return any(not i.get("senha_hash") for i in carregar_usuarios().values())


def senha_unica():
    """Senha compartilhada definida nos Secrets do Streamlit (para repositorio publico).
    Se estiver definida, o app usa login so com senha e ignora o users.yaml."""
    try:
        if "APP_PASSWORD" in st.secrets:
            return str(st.secrets["APP_PASSWORD"])
    except Exception:
        pass
    return os.environ.get("APP_PASSWORD", "")


def tela_login_simples():
    st.title("📄 Conversor de extratos")
    st.caption("Ferramentas internas — acesso restrito")
    with st.form("login_s"):
        senha = st.text_input("Senha de acesso", type="password")
        if st.form_submit_button("Entrar", type="primary", use_container_width=True):
            if senha and senha == senha_unica():
                st.session_state.logado = True
                st.session_state.usuario = "equipe"
                st.session_state.usuario_info = {"nome": "Equipe", "perfil": "padrao"}
                st.rerun()
            else:
                st.error("Senha incorreta.")


def tela_login():
    st.title("📄 Conversor de extratos")
    st.caption("Ferramentas internas — acesso restrito")
    with st.form("login"):
        usuario = st.text_input("Usuário", placeholder="bruna")
        senha   = st.text_input("Senha", type="password")
        if st.form_submit_button("Entrar", type="primary", use_container_width=True):
            ok, info = verificar_login(usuario, senha)
            if ok:
                st.session_state.logado = True
                st.session_state.usuario = usuario.lower().strip()
                st.session_state.usuario_info = info
                st.rerun()
            else:
                st.error("Usuário ou senha inválidos.")


def tela_configurar_senha():
    st.title("Primeiro acesso — defina sua senha")
    usuarios = carregar_usuarios()
    sem = [(u, i) for u, i in usuarios.items() if not i.get("senha_hash")]
    for u, i in sem:
        with st.form(f"cfg_{u}"):
            st.write(f"**{i.get('nome', u)}** (`{u}`)")
            s1 = st.text_input("Senha", type="password", key=f"s1_{u}")
            s2 = st.text_input("Confirmar senha", type="password", key=f"s2_{u}")
            if st.form_submit_button("Salvar"):
                if len(s1) < 4:
                    st.error("Mínimo de 4 caracteres.")
                elif s1 != s2:
                    st.error("As senhas não conferem.")
                else:
                    usuarios[u]["senha_hash"] = criar_hash(s1)
                    salvar_usuarios(usuarios)
                    st.success("Senha salva! Recarregue e faça login.")


# ============================================================================
# 2. LEITURA DE OFX  (deterministica)
# ============================================================================
def _tag(bloco, tag):
    """Le o valor de uma tag OFX (SGML ou XML). Vai ate a proxima tag ou quebra de linha."""
    m = re.search(rf"<{tag}>([^<\r\n]*)", bloco, re.IGNORECASE)
    return m.group(1).strip() if m else ""


def _data_ofx(bruto):
    """20260701  ou  20260701120000[-3:GMT]  ->  01/07/2026"""
    m = re.match(r"\s*(\d{4})(\d{2})(\d{2})", bruto or "")
    if not m:
        return ""
    a, mes, d = m.groups()
    return f"{d}/{mes}/{a}"


def ler_ofx(conteudo_bytes):
    texto = None
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            texto = conteudo_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        texto = conteudo_bytes.decode("latin-1", errors="ignore")

    blocos = re.findall(r"<STMTTRN>(.*?)</STMTTRN>", texto, re.IGNORECASE | re.DOTALL)
    if not blocos:
        # alguns OFX SGML nao fecham a tag; separa por abertura
        blocos = re.split(r"<STMTTRN>", texto, flags=re.IGNORECASE)[1:]

    transacoes = []
    for b in blocos:
        valor_txt = _tag(b, "TRNAMT")
        if "," in valor_txt:                       # OFX fora do padrao (virgula decimal)
            valor_txt = valor_txt.replace(".", "").replace(",", ".")
        try:
            valor = float(valor_txt)
        except ValueError:
            continue
        data = _data_ofx(_tag(b, "DTPOSTED") or _tag(b, "DTUSER"))
        desc = _tag(b, "MEMO") or _tag(b, "NAME") or _tag(b, "PAYEE")
        desc = re.sub(r"\s+", " ", desc).strip()
        trntype = _tag(b, "TRNTYPE").upper()

        if valor < 0 or trntype in ("DEBIT", "PAYMENT", "FEE", "SRVCHG", "ATM", "CASH", "DIRECTDEBIT", "CHECK"):
            tipo = "debito"
        elif valor > 0 or trntype in ("CREDIT", "DEP", "DIRECTDEP", "INT", "XFER"):
            tipo = "credito"
        else:
            tipo = "debito" if valor < 0 else "credito"

        if not data or valor == 0:
            continue
        transacoes.append({
            "data": data,
            "descricao": desc or trntype or "(sem descrição)",
            "valor": round(abs(valor), 2),
            "tipo": tipo,
        })
    return transacoes


# ============================================================================
# 3. LEITURA DE PDF via API da Anthropic (Claude)
# ============================================================================
PROMPT_EXTRACAO = """Você é um extrator de dados de extratos bancários brasileiros.

Receberá as páginas de um extrato bancário (qualquer banco). Extraia TODAS as
movimentações (lançamentos) e devolva SOMENTE um JSON válido, sem texto antes ou
depois, no formato:

{"transacoes": [
  {"data": "DD/MM/AAAA", "descricao": "texto do histórico", "valor": 1234.56, "tipo": "debito"}
]}

Regras:
- "data": sempre no formato DD/MM/AAAA. Se o ano não aparecer na linha, use o ano do período do extrato.
- "valor": número positivo, ponto como separador decimal, sem separador de milhar, sem "R$".
- "tipo": "debito" para dinheiro que SAI da conta (saída, débito, pagamento, saque, tarifa, compra, PIX enviado, IOF, juros pagos);
          "credito" para dinheiro que ENTRA (entrada, crédito, depósito, recebimento, PIX recebido, rendimento, estorno a favor).
- Use o sinal (- ou +) ou a marcação D/C da linha quando existir para decidir o tipo.
- Junte descrições que continuam na linha seguinte.
- Mantenha a ordem em que aparecem no extrato.
- IGNORE linhas que não são movimentação: "SALDO ANTERIOR", "SALDO DO DIA", "SALDO ATUAL",
  "SALDO BLOQUEADO", "S A L D O", subtotais, totais, saldos de fechamento, cabeçalhos e rodapés de página.
- Não invente lançamentos. Se o valor de uma linha estiver ilegível, pule a linha.
"""


def _paginas_texto(pdf_bytes):
    """Extrai o texto de cada pagina. Retorna lista de strings (ou [] se nao der)."""
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(pdf_bytes))
        return [(p.extract_text() or "") for p in reader.pages]
    except Exception:
        return []


def _chunks_pdf_bytes(pdf_bytes, n_por_chunk):
    """Divide o PDF em PDFs menores. Retorna lista de (rotulo, bytes)."""
    try:
        from pypdf import PdfReader, PdfWriter
        reader = PdfReader(io.BytesIO(pdf_bytes))
    except Exception:
        return [("arquivo inteiro", pdf_bytes)]
    n = len(reader.pages)
    out = []
    for ini in range(0, n, n_por_chunk):
        fim = min(ini + n_por_chunk, n)
        w = PdfWriter()
        for i in range(ini, fim):
            w.add_page(reader.pages[i])
        buf = io.BytesIO()
        w.write(buf)
        out.append((f"páginas {ini + 1}-{fim}", buf.getvalue()))
    return out or [("arquivo inteiro", pdf_bytes)]


def _extrair_json(texto):
    texto = texto.strip()
    texto = re.sub(r"^```(?:json)?\s*", "", texto)
    texto = re.sub(r"\s*```$", "", texto).strip()
    try:
        return json.loads(texto)
    except Exception:
        pass
    m = re.search(r"\{.*\}", texto, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(0))
        except Exception:
            pass
    # salvamento: recupera objetos soltos de uma resposta truncada
    objs = []
    for pedaco in re.findall(r"\{[^{}]*\}", texto, re.DOTALL):
        try:
            o = json.loads(pedaco)
            if "valor" in o or "data" in o:
                objs.append(o)
        except Exception:
            continue
    return {"transacoes": objs}


def _get_api_key():
    try:
        if "ANTHROPIC_API_KEY" in st.secrets:
            return str(st.secrets["ANTHROPIC_API_KEY"]).strip()
    except Exception:
        pass
    return os.environ.get("ANTHROPIC_API_KEY", "").strip()


def _lote_claude(client, blocks, rotulo):
    """Chama a IA para um lote (com streaming e re-tentativas).
    Retorna (transacoes, observacao_de_erro)."""
    ultimo = ""
    for tentativa in range(1, TENTATIVAS_LOTE + 1):
        try:
            with client.messages.stream(
                model=MODELO_CLAUDE,
                max_tokens=MAX_TOKENS_SAIDA,
                messages=[{"role": "user", "content": blocks}],
            ) as stream:
                msg = stream.get_final_message()
            txt = "".join(b.text for b in msg.content if getattr(b, "type", "") == "text")
            dados = _extrair_json(txt)
            trans = dados.get("transacoes", []) if isinstance(dados, dict) else []
            if msg.stop_reason == "max_tokens":
                return trans, (f"{rotulo}: resposta muito longa — parte dos lançamentos "
                               f"pode ter ficado de fora (diminua CHUNK_PAGINAS_TEXTO no app).")
            return trans, ""
        except Exception as e:
            ultimo = (str(e).splitlines() or [repr(e)])[0][:180]
            time.sleep(2 * tentativa)
    return [], f"{rotulo}: falhou após {TENTATIVAS_LOTE} tentativas — {ultimo}"


def ler_pdf(pdf_bytes, progresso=None):
    api_key = _get_api_key()
    if not api_key:
        raise RuntimeError(
            "Chave da API da Anthropic não configurada. "
            "Adicione ANTHROPIC_API_KEY nos Secrets do app (veja o README)."
        )
    import anthropic
    client = anthropic.Anthropic(api_key=api_key, timeout=600.0)

    def _rodar_modo_texto(paginas):
        n = len(paginas)
        lotes = [(i, min(i + CHUNK_PAGINAS_TEXTO, n)) for i in range(0, n, CHUNK_PAGINAS_TEXTO)]
        todas, falhas = [], []
        for k, (a, b) in enumerate(lotes, 1):
            if progresso:
                progresso(k, len(lotes))
            corpo = "\n\n".join(f"[página {a + j + 1}]\n{paginas[a + j]}" for j in range(b - a))
            blocks = [{"type": "text", "text": PROMPT_EXTRACAO + "\n\n=== EXTRATO ===\n" + corpo}]
            trans, obs = _lote_claude(client, blocks, f"páginas {a + 1}-{b}")
            todas += trans
            if obs:
                falhas.append(obs)
        return todas, falhas

    def _rodar_modo_pdf(pdf_bytes):
        lotes = _chunks_pdf_bytes(pdf_bytes, CHUNK_PAGINAS_PDF)
        todas, falhas = [], []
        for k, (rot, by) in enumerate(lotes, 1):
            if progresso:
                progresso(k, len(lotes))
            b64 = base64.standard_b64encode(by).decode()
            blocks = [
                {"type": "document",
                 "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
                {"type": "text", "text": PROMPT_EXTRACAO},
            ]
            trans, obs = _lote_claude(client, blocks, rot)
            todas += trans
            if obs:
                falhas.append(obs)
        return todas, falhas

    paginas = _paginas_texto(pdf_bytes)
    media_chars = (sum(len(p.strip()) for p in paginas) / len(paginas)) if paginas else 0

    if media_chars > 120:                       # PDF com texto (rápido e barato)
        todas, falhas = _rodar_modo_texto(paginas)
        if not todas:                           # deu ruim -> tenta como PDF escaneado
            todas, falhas = _rodar_modo_pdf(pdf_bytes)
    else:                                        # PDF escaneado (imagem)
        todas, falhas = _rodar_modo_pdf(pdf_bytes)

    avisos = []
    if falhas:
        avisos.append("⚠️ Nem todo o extrato foi lido — os lançamentos abaixo podem estar "
                      "INCOMPLETOS. Trechos com problema:\n\n- " + "\n- ".join(falhas)
                      + "\n\nTente enviar de novo (às vezes resolve) ou divida o PDF em partes menores.")

    # remove duplicata exata só na fronteira entre lotes (lançamento cortado pela quebra de página)
    limpo = []
    for t in todas:
        chave = (t.get("data"), t.get("descricao"), t.get("valor"), t.get("tipo"))
        if limpo and chave == (limpo[-1].get("data"), limpo[-1].get("descricao"),
                               limpo[-1].get("valor"), limpo[-1].get("tipo")):
            continue
        limpo.append(t)
    return limpo, avisos


# ============================================================================
# 4. NORMALIZACAO + GERACAO DO EXCEL
# ============================================================================
def normalizar(transacoes):
    """Padroniza data, valor e tipo. Descarta linhas invalidas."""
    out = []
    for t in transacoes:
        data = str(t.get("data", "")).strip()
        m = re.search(r"(\d{1,2})[/.\-](\d{1,2})[/.\-](\d{2,4})", data)
        if m:
            d, mes, a = m.groups()
            if len(a) == 2:
                a = "20" + a
            data = f"{int(d):02d}/{int(mes):02d}/{a}"
        else:
            continue

        valor = t.get("valor", "")
        if isinstance(valor, str):
            v = valor.replace("R$", "").strip()
            if "," in v and "." in v:
                v = v.replace(".", "").replace(",", ".")
            elif "," in v:
                v = v.replace(",", ".")
            try:
                valor = float(v)
            except ValueError:
                continue
        try:
            valor = round(abs(float(valor)), 2)
        except (TypeError, ValueError):
            continue
        if valor == 0:
            continue

        tipo = str(t.get("tipo", "")).lower()
        tipo = "credito" if tipo.startswith("cred") else "debito"

        desc = re.sub(r"\s+", " ", str(t.get("descricao", ""))).strip() or "(sem descrição)"
        out.append({"data": data, "descricao": desc, "valor": valor, "tipo": tipo})
    return out


def _chave_data(d):
    try:
        return datetime.strptime(d, "%d/%m/%Y")
    except ValueError:
        return datetime.max


def gerar_excel(transacoes):
    debitos  = sorted([t for t in transacoes if t["tipo"] == "debito"],  key=lambda t: _chave_data(t["data"]))
    creditos = sorted([t for t in transacoes if t["tipo"] == "credito"], key=lambda t: _chave_data(t["data"]))

    wb = Workbook()
    ws_d = wb.active
    ws_d.title = "Debito"
    ws_c = wb.create_sheet("Credito")

    for ws, linhas in ((ws_d, debitos), (ws_c, creditos)):
        ws.append(CABECALHO)
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
        for t in linhas:
            ws.append([t["data"], t["descricao"], t["valor"], ""])
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 20
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            row[0].number_format = "#,##0.00"
        ws.auto_filter.ref = f"A1:D{ws.max_row}"   # filtro na linha 1
        ws.freeze_panes = "A2"                      # congela a linha 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), len(debitos), len(creditos)


# ============================================================================
# 5. TELA PRINCIPAL
# ============================================================================
def tela_principal():
    col1, col2 = st.columns([6, 1])
    with col1:
        st.title("📄 Conversor de extratos → Excel")
    with col2:
        if st.button("Sair"):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()

    st.caption("Anexe o extrato em **PDF** ou **OFX**. O resultado sai no padrão "
               "de duas abas: **Debito** e **Credito** (Data | Descrição | Valor | Obs).")

    arquivo = st.file_uploader("Extrato bancário", type=["pdf", "ofx"])

    if arquivo is None:
        st.info("Aguardando arquivo…")
        return

    ext = arquivo.name.lower().rsplit(".", 1)[-1]
    dados_brutos = arquivo.getvalue()

    # ---- processa (com cache por conteúdo) ----
    assinatura = hashlib.md5(dados_brutos).hexdigest()
    if st.session_state.get("assinatura") != assinatura:
        try:
            if ext == "ofx":
                with st.spinner("Lendo OFX…"):
                    transacoes = ler_ofx(dados_brutos)
                avisos = []
            else:
                barra = st.progress(0.0, text="Lendo o PDF…")
                def _p(i, n):
                    barra.progress(i / n, text=f"Lendo o PDF… parte {i} de {n}")
                transacoes, avisos = ler_pdf(dados_brutos, progresso=_p)
                barra.empty()
        except Exception as e:
            st.error(f"Falha ao ler o arquivo: {e}")
            return

        transacoes = normalizar(transacoes)
        st.session_state.assinatura = assinatura
        st.session_state.transacoes = transacoes
        st.session_state.avisos = avisos

    transacoes = st.session_state.get("transacoes", [])
    for a in st.session_state.get("avisos", []):
        (st.error if a.startswith("⚠️") else st.warning)(a)

    if not transacoes:
        st.error("Nenhuma movimentação foi identificada. "
                 "Se o PDF for escaneado, confira se está legível.")
        return

    # ---- revisão / edição ----
    st.subheader("Confira e ajuste antes de exportar")
    try:
        import pandas as pd
        df = pd.DataFrame(transacoes, columns=["data", "descricao", "valor", "tipo"])
        df_edit = st.data_editor(
            df,
            use_container_width=True,
            num_rows="dynamic",
            column_config={
                "data": st.column_config.TextColumn("Data", width="small"),
                "descricao": st.column_config.TextColumn("Descrição", width="large"),
                "valor": st.column_config.NumberColumn("Valor", format="%.2f"),
                "tipo": st.column_config.SelectboxColumn("Tipo", options=["debito", "credito"]),
            },
            key="editor",
        )
        transacoes_final = normalizar(df_edit.to_dict("records"))
    except Exception:
        st.dataframe(transacoes, use_container_width=True)
        transacoes_final = transacoes

    tot_d = sum(t["valor"] for t in transacoes_final if t["tipo"] == "debito")
    tot_c = sum(t["valor"] for t in transacoes_final if t["tipo"] == "credito")
    c1, c2, c3 = st.columns(3)
    c1.metric("Lançamentos", len(transacoes_final))
    c2.metric("Total débitos",  f"R$ {tot_d:,.2f}")
    c3.metric("Total créditos", f"R$ {tot_c:,.2f}")

    # ---- exporta ----
    nome_base = re.sub(r"\.(pdf|ofx)$", "", arquivo.name, flags=re.IGNORECASE)
    xlsx, nd, nc = gerar_excel(transacoes_final)
    st.download_button(
        f"⬇ Baixar Excel  ({nd} débitos · {nc} créditos)",
        data=xlsx,
        file_name=f"{nome_base}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )


# ============================================================================
# Roteamento
# ============================================================================
if "logado" not in st.session_state:
    st.session_state.logado = False

if senha_unica():
    # modo senha unica (Secrets) -> serve para repositorio publico
    if not st.session_state.logado:
        tela_login_simples()
    else:
        tela_principal()
else:
    # modo multiusuario (users.yaml) -> repositorio privado
    if not carregar_usuarios():
        st.error("Configure o acesso: defina `APP_PASSWORD` nos Secrets "
                 "ou inclua um `users.yaml` no projeto.")
    elif precisa_configurar_senha() and not st.session_state.logado:
        tela_configurar_senha()
        st.stop()
    elif not st.session_state.logado:
        tela_login()
    else:
        tela_principal()
